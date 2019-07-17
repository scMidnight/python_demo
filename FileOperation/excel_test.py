#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
#引用openpyxl 。

#写入Excel

wb = openpyxl.Workbook()
#利用openpyxl.Workbook()函数创建新的workbook（工作薄）对象，就是创建新的空的Excel文件。

sheet = wb.active
#wb.active就是获取这个工作薄的活动表，通常就是第一个工作表。

sheet.title = 'new title'
#可以用.title给工作表重命名。现在第一个工作表的名称就会由原来默认的“sheet1”改为"new title"。

sheet['A1'] = '漫威宇宙'
#把'漫威宇宙'赋值给第一个工作表的A1单元格，就是往A1的单元格中写入了'漫威宇宙'。

row = ['美国队长','钢铁侠','蜘蛛侠']
#把我们想写入的一行内容写成列表，赋值给row。

sheet.append(row)
#用sheet.append()就能往表格里添加这一行文字。

rows = [['美国队长','钢铁侠','蜘蛛侠'],['是','漫威','宇宙', '经典','人物']]
#先把要写入的多行内容写成列表，再放进大列表里，赋值给rows。

for i in rows:
    sheet.append(i)
#遍历rows，同时把遍历的内容添加到表格里，这样就实现了多行写入。

print(rows)
#打印rows

wb.save('Marvel.xlsx')
#保存新建的Excel文件，并命名为“Marvel.xlsx”


#读取excel文件

wb = openpyxl.load_workbook('Marvel.xlsx')
sheet = wb['new title']
sheetname = wb.sheetnames
print(sheetname)
A1_cell = sheet['A1']
A1_value = A1_cell.value
print(A1_value)
print(sheet.max_row)
print(sheet.max_column)
for row in sheet.rows:
    for cell in row:
        print(cell.value)
